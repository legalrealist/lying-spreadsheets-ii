"""
Defense for the cached-value attack: recompute every formula from its raw
precedents and diff against the stored cache. Any cell where the recomputed
result != the cached <v> is a tampered (or merely stale) cell.

This module ships a small, dependency-light recompute engine that resolves cell
references bottom-up from non-formula inputs across ALL sheets (it follows
`Sheet!Ref` cross-sheet references), so a tampered cache on a formula cell does
not poison the recomputation. It supports the common operators and
IF/SUM/MIN/MAX/ROUND/ABS/AVERAGE -- enough for the included demos and many
linear models. It does NOT handle ranges, quoted sheet names with spaces, or
external-workbook links; for arbitrary workbooks use a full engine:

    * LibreOffice headless: soffice --headless --calc --convert-to xlsx (forces recalc)
    * pip install formulas   (pure-python Excel calculator)
    * pip install pycel

Crucially, recompute can only catch a tamper when the precedents are present. If
the pipeline extracted only a summary sheet and dropped the inputs it references,
the formula's precedent is gone and this detector returns nothing for that cell
(it cannot verify) -- the same blind spot non-numeric data has. See
experiments/nonnumeric.md.

Usage:
    python defense/detect_xlsx.py path/to/workbook.xlsx
"""
from __future__ import annotations

import re
import sys

import openpyxl

_FUNCS = {"IF": "_if", "SUM": "_sum", "MIN": "min", "MAX": "max",
          "ROUND": "_round", "ABS": "abs", "AVERAGE": "_avg"}


def _if(cond, a, b):
    return a if cond else b


def _sum(*args):
    return sum(_flatten(args))


def _avg(*args):
    vals = _flatten(args)
    return sum(vals) / len(vals) if vals else 0


def _round(x, n=0):
    return round(x, int(n))


def _flatten(args):
    out = []
    for a in args:
        out.extend(a) if isinstance(a, (list, tuple)) else out.append(a)
    return [v for v in out if isinstance(v, (int, float))]


_NS = {"_if": _if, "_sum": _sum, "_avg": _avg, "_round": _round,
       "min": min, "max": max, "abs": abs}

_XREF = re.compile(r"([A-Za-z_][A-Za-z0-9_]*)!\$?([A-Z]{1,3})\$?(\d+)")
_BARE = re.compile(r"(?<![A-Za-z0-9_!])\$?([A-Z]{1,3})\$?(\d+)\b")


def _eval(formula: str, sheet: str, resolve) -> object:
    """Evaluate one Excel formula by substituting resolved precedent values."""
    f = formula.lstrip("=")
    strings = []
    f = re.sub(r'"[^"]*"', lambda m: (strings.append(m.group(0)), f"\x01{len(strings) - 1}\x01")[1], f)
    for name, fn in _FUNCS.items():
        f = re.sub(rf"\b{name}\s*\(", fn + "(", f, flags=re.I)
    f = f.replace("<>", "!=")
    f = re.sub(r"(?<![<>=!])=(?!=)", "==", f)
    f = _XREF.sub(lambda m: repr(resolve(f"{m.group(1)}!{m.group(2)}{m.group(3)}")), f)
    f = _BARE.sub(lambda m: repr(resolve(f"{sheet}!{m.group(1)}{m.group(2)}")), f)
    for i, s in enumerate(strings):
        f = f.replace(f"\x01{i}\x01", s)
    return eval(f, {"__builtins__": {}}, _NS)  # noqa: S307 (no refs/builtins in scope)


def recompute_all(path: str) -> dict:
    """Return {`Sheet!Ref`: recomputed_value} for every formula cell, workbook-wide."""
    wb = openpyxl.load_workbook(path, data_only=False)
    inputs, formulas = {}, {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                key = f"{ws.title}!{cell.coordinate}"
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas[key] = (ws.title, cell.value)
                else:
                    inputs[key] = cell.value

    memo: dict = {}

    def resolve(key: str):
        if key in memo:
            return memo[key]
        if key in inputs:
            memo[key] = inputs[key]
        elif key in formulas:
            sheet, f = formulas[key]
            try:
                memo[key] = _eval(f, sheet, resolve)
            except Exception:
                memo[key] = None  # precedent missing / unsupported -> cannot verify
        else:
            memo[key] = None      # referenced precedent not present in the workbook
        return memo[key]

    return {k: resolve(k) for k in formulas}


def detect(path: str, tol: float = 1e-6) -> list:
    """Return [(`Sheet!Ref`, cached, recomputed)] where cache != recomputation.

    Cells whose precedents are absent recompute to None and are skipped (the
    detector cannot verify them) rather than reported as clean.
    """
    wb_c = openpyxl.load_workbook(path, data_only=True)
    cached = {f"{ws.title}!{c.coordinate}": c.value
              for ws in wb_c.worksheets for row in ws.iter_rows() for c in row}
    findings = []
    for key, recomputed in recompute_all(path).items():
        cv = cached.get(key)
        if cv is None or recomputed is None:
            continue
        if isinstance(recomputed, (int, float)) and isinstance(cv, (int, float)):
            mismatch = abs(float(recomputed) - float(cv)) > tol
        else:
            mismatch = str(recomputed) != str(cv)
        if mismatch:
            findings.append((key, cv, recomputed))
    return findings


def main() -> int:
    if len(sys.argv) != 2:
        print(__doc__)
        return 2
    findings = detect(sys.argv[1])
    if not findings:
        print("OK -- no cache/recomputation divergence found (note: cells with absent "
              "precedents cannot be verified).")
        return 0
    print(f"TAMPERING SUSPECTED -- {len(findings)} cell(s) where cache != recomputation:")
    for ref, cached, recomputed in findings:
        print(f"  {ref}: cached={cached!r}  recomputed={recomputed!r}")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
