"""
Non-numeric cached-value spoofing -- the spreadsheet analog of the E3 email case.

A "Deal Summary" sheet shows text fields (Governing Law, Counterparty Rating,
Maturity) that are *formulas* pulling from a separate "Refs" inputs sheet. Their
cached values are tampered. There is no arithmetic to cross-check.

    human (Excel)     -> recalculates =Refs!B1 -> TRUE text ("New York", "BBB-")
    pipeline (pandas) -> reads the fabricated cache -> ATTACKER text ("Delaware", "AAA")

The recompute defense still catches this *when the Refs sheet is present* (it
resolves the cross-sheet reference). But pipelines often extract a single named
sheet; drop Refs and the precedent is gone, so recompute goes dark -- exactly
like a non-numeric field with nothing to check. See experiments/nonnumeric.md.

Run: python poc/lying_xlsx_text.py
"""
from __future__ import annotations

import os
import sys

import openpyxl
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from poc.lying_xlsx import tamper  # noqa: E402

SUMMARY_ROWS = [
    ("Governing Law", "=Refs!B1"),       # B1  true "New York"
    ("Counterparty Rating", "=Refs!B2"),  # B2  true "BBB-"
    ("Maturity", "=Refs!B3"),             # B3  true "2030-12-31"
]
REFS = {"B1": "New York", "B2": "BBB-", "B3": "2030-12-31"}

# tamper the Deal Summary cached text to favorable values (all text -> is_text=True)
TAMPER = {"B1": ("Delaware", True), "B2": ("AAA", True), "B3": ("2035-12-31", True)}


def build_deal(path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DealSummary"
    for i, (label, formula) in enumerate(SUMMARY_ROWS, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=formula)
    refs = wb.create_sheet("Refs")
    for ref, val in REFS.items():
        refs[ref] = val
    wb.save(path)
    return path


def make_summary_only(src: str, dst: str) -> str:
    """Simulate a pipeline that extracted only the Deal Summary sheet."""
    wb = openpyxl.load_workbook(src)
    if "Refs" in wb.sheetnames:
        del wb["Refs"]
    wb.save(dst)
    return dst


def pandas_read(path: str, sheet: str = "DealSummary") -> dict:
    df = pd.read_excel(path, sheet_name=sheet, header=None, usecols=[0, 1], names=["Item", "Value"])
    return {str(r.Item): r.Value for r in df.itertuples()}


def main() -> None:
    build_deal("/tmp/_deal.xlsx")
    tamper("/tmp/_deal.xlsx", "/tmp/_deal_t.xlsx", TAMPER)
    print("pipeline (pandas) reads:", pandas_read("/tmp/_deal_t.xlsx"))
    print("human (Excel recalc)   :", {"Governing Law": REFS["B1"], "Counterparty Rating": REFS["B2"], "Maturity": REFS["B3"]})
    for f in ("/tmp/_deal.xlsx", "/tmp/_deal_t.xlsx"):
        if os.path.exists(f):
            os.remove(f)


if __name__ == "__main__":
    main()
